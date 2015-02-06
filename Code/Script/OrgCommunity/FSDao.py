#coding=utf-8
import pickle
from igraph import *
import xlwt
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
import os

__author__ = 'Administrator'

'''
Zuo:
In this module I am going to write some code to access the graph from the file system
finished:

todo:
make graph from the raw csv file
serialize the graph to the file system with pajek(or other formal) format
read the graph from the fs which the format is pajek(or other formal) format
'''



def make_graph(csv_node_file,edge_node_file):
    '''
    Zuo:
    Implement how we convert the raw file(retrieve from sql server) to the graph object we need
    '''

    g=Graph(n=0,directed=False)

    #begin to add vertices
    #Notice that the igraph index differs from the dbIndex since they start from 0 and 1 respectively
    f1=open(csv_node_file,'r')
    for line in f1:
        l=line.strip('﻿ ').split(',')
        g.add_vertex(name=l[1],dbIndex=int(l[0]),occur_count=int(l[2]))
    f1.close()

    print 'add %s vertices'%g.vcount()

    # begin to add edges
    f2=open(edge_node_file)
    edge_list=[]
    edge_attr_list=[]
    for line in f2:
        l=line.strip('﻿ ').split(',')
        edge_list.append((int(l[0])-1,int(l[1])-1))
        edge_attr_list.append(int(l[2]))
    f2.close()
    g.add_edges(edge_list)
    g.es['co_occur']=edge_attr_list
    print 'add %s edges'%g.ecount()
    return g


def write_graph(g,filepath):
    '''
    This is the interface for writing the graph into the given format
    '''
    f=open(filepath,'w')

    #can pajek store the information of the graph?
    if isinstance(g,Graph):
        g.write_pickle(f)
    else:
        print 'The object going to write is not a graph'
    f.close()


def read_pickle_graph(filename):
    '''
    Actually I am not quite sure what this for, it doesnt work for community_extraction_simplerelation
    '''

    if os.path.exists(filename):
        import igraph
        g=Graph.Read_Pickle(filename)
        return g
    else:
        print '%s not exists'%filename

def write_pickle(directory, filename, obj):
    '''
    write the pickle to the given directory and filename
    '''
    if not os.path.exists(directory):
        os.makedirs(directory)
    f=open(directory+filename,'wb')
    pickle.dump(obj,f)
    f.close()

def append_excel_sheet(directory,filename, sheetname, headings, data, comment=None):
    '''
    append a sheet to a existing excel file
    if the file does not exist, create it
    '''
    pass

def write_excel(directory,filename, sheetname,headings, data,comment=None):
    wb=Workbook()
    ew=ExcelWriter(workbook=wb)
    ws=wb.worksheets[0]
    ws.title=sheetname
    #begin to write the headings
    row=1
    for i,heading in enumerate(headings):
        col=get_column_letter(i+1)
        ws.cell('%s%s'%(col,row)).value='%s'%(heading)

    #write the cell content
    for row,rowdata in enumerate(data):
        for i,celldataKey in enumerate(rowdata):
            col=get_column_letter(i+1)
            ws.cell('%s%s'%(col,(row+2))).value='%s'%(rowdata[celldataKey])

    if not os.path.exists(directory):
        os.makedirs(directory)
    ew.save(directory+filename)

from openpyxl.reader.excel import load_workbook
def read_excel(directory,filename,sheetname):

    wb=load_workbook(directory+filename)
    ws=wb.get_sheet_by_name(sheetname)

    # read heading
    cols=ws.get_highest_column()
    headings=[]
    for i in range(cols):
        headings.append(ws.cell(row=0,column=i).value)

    info=[]
    for rx in range(1,ws.get_highest_row()):
        d={}
        for col,k in enumerate(headings):
            d[k]=ws.cell(row=rx,column=col).value
        info.append(d)

    return info

if __name__=='__main__1':
    g=make_graph('E:\\Graduation\\Data\\nodes_utf.csv', 'E:\\Graduation\\Data\\edges_utf.csv')
    g.write_pickle('E:/Graduation/Data/word.pickle')

#read pickle from the file system
if __name__=='__main__':
    g=read_pickle_graph('E:/Graduation/Data/word.pickle')
    print g.summary()
