#coding=utf-8
__author__ = 'Administrator'

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

def write_excel(filename, sheetname,headings, data):
    wb=Workbook()
    ew=ExcelWriter(workbook=wb)
    ws=wb.worksheets[0]
    ws.title=sheetname
    #begin to write the headings
    row=1
    for i,heading in enumerate(headings):
        col=get_column_letter(i+1)
        ws.cell('%s%s'%(col,row)).value='%s'%(heading)

    for row,rowdata in enumerate(data):
        for i,celldata in enumerate(rowdata):
            col=get_column_letter(i+1)
            ws.cell('%s%s'%(col,(row+2))).value='%s'%(celldata)
    ew.save(filename)

if __name__=='__main__':
    headings=['word_id','noun','group_id']
    data=[[1,'微博',3]]*70000
    write_excel('test_file.xlsx','1',headings,data)